# =============================================================================
# Nested CV with PyKEEN graph embeddings (TransE, HoLE)
# - Shared cached train/test/validation split across both embeddings
# - Cached trained embedding models (skip retraining if already saved)
# - Cached CUI -> Neo4j entity ID lookups (avoid repeated DB queries)
# - Outer 10-fold CV, Inner 5-fold GridSearchCV (same grid as RF experiment)
# - Class 0 (majority) is RandomUnderSampler-downsampled by divisor in {5, 10}
#   Classes 1-5 (minority) are SMOTE-upsampled by multiplier in {2, 3}
#   -> 4 combos: (5,2), (5,3), (10,2), (10,3)
# - Detailed per-fold, per-class Excel export, one file per embedding
# =============================================================================

import os
import pickle
import itertools
import numpy as np
import pandas as pd
import torch
from torch.autograd import Variable
from collections import Counter
from typing import List

from neo4j import GraphDatabase
from pykeen.triples import TriplesFactory
from pykeen.pipeline import pipeline
import pykeen.nn

from sklearn.ensemble import RandomForestClassifier
from sklearn.feature_selection import SelectFromModel
from sklearn.model_selection import StratifiedKFold, GridSearchCV
from sklearn.pipeline import Pipeline
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.utils import resample
from imblearn.over_sampling import SMOTE
from imblearn.under_sampling import RandomUnderSampler
from imblearn.pipeline import Pipeline as ImbPipeline

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIG
# =============================================================================

EMBEDDINGS_TO_RUN = ['TransE', 'HoLE']
DEVICE = 'cuda:2'

SPLIT_CACHE_DIR = 'split_cache'
CUI_ID_CACHE_PATH = 'cui_id_cache.pkl'

GROUND_TRUTH_PATH = "/home/faisopos/workspace/marios/CUIs_new.csv"

param_grid_pipeline = {
    'clf__n_estimators': [100, 200, 500],
    'clf__max_depth': [5, 10],
    'clf__min_samples_leaf': [5, 10],
}

# Class 0 (majority) is downsampled by dividing its count by `divisor`.
# Classes 1-5 (minority) are SMOTE-upsampled by `multiplier`.
downsample_divisors = [5, 10]
upsample_multipliers = [2, 3]
sampling_combos = list(itertools.product(downsample_divisors, upsample_multipliers))
# -> [(5, 2), (5, 3), (10, 2), (10, 3)]

OUTER_K = 10
INNER_K = 5


# =============================================================================
# Neo4j connection
# =============================================================================

print('Neo4j connection details...')
host = 'bolt://localhost:7687'
user = 'neo4j'
password = 'iasis'
driver = GraphDatabase.driver(host, auth=(user, password))


def run_query(query, params={}):
    with driver.session() as session:
        result = session.run(query, params)
        return pd.DataFrame([r.values() for r in result], columns=result.keys())


# --- CUI -> Neo4j internal ID cache (avoids re-querying Neo4j for repeats) ---
if os.path.exists(CUI_ID_CACHE_PATH):
    with open(CUI_ID_CACHE_PATH, 'rb') as f:
        cui_id_cache = pickle.load(f)
    print(f"Loaded CUI->ID cache with {len(cui_id_cache)} entries.")
else:
    cui_id_cache = {}


def getIdFromCUI(cui):
    if cui in cui_id_cache:
        return cui_id_cache[cui]
    cquery = r"MATCH (s:Entity) WHERE (s.id='" + cui + r"') RETURN toString(id(s)) as id"
    loc_id = run_query(cquery)
    if len(loc_id) == 0:
        cui_id_cache[cui] = "null"
        return "null"
    result_id = loc_id['id'][0]
    cui_id_cache[cui] = result_id
    return result_id


def save_cui_cache():
    with open(CUI_ID_CACHE_PATH, 'wb') as f:
        pickle.dump(cui_id_cache, f)


# =============================================================================
# Shared triples factory + train/test/validation split (cached once, reused
# by both embeddings so they see the IDENTICAL split)
# =============================================================================

os.makedirs(SPLIT_CACHE_DIR, exist_ok=True)
labeled_triples_path = os.path.join(SPLIT_CACHE_DIR, 'labeled_triples.pkl')
splits_path = os.path.join(SPLIT_CACHE_DIR, 'splits.pkl')

if os.path.exists(labeled_triples_path) and os.path.exists(splits_path):
    print('Found cached triples + split. Loading from disk (skipping Neo4j query)...')

    with open(labeled_triples_path, 'rb') as f:
        labeled_triples = pickle.load(f)

    # Rebuild the triples factory exactly as it was built the first time, so
    # entity_to_id / relation_to_id mappings come out identical.
    tf = TriplesFactory.from_labeled_triples(
        labeled_triples,
        create_inverse_triples=False,
        entity_to_id=None,
        relation_to_id=None,
        compact_id=False,
        filter_out_candidate_inverse_relations=True,
        metadata=None,
    )

    with open(splits_path, 'rb') as f:
        split_indices = pickle.load(f)
    training = tf.clone_and_exchange_triples(split_indices['train'])
    testing = tf.clone_and_exchange_triples(split_indices['test'])
    validation = tf.clone_and_exchange_triples(split_indices['valid'])
    print('Done loading cached split.')
else:
    print('No cache found. Querying Neo4j for full graph...')
    data = run_query("""
    MATCH (s)-[r]->(t)
    RETURN toString(id(s)) as source, toString(id(t)) AS target, type(r) as type
    """)

    labeled_triples = data[["source", "type", "target"]].values

    print('Insert neo4j graph into pykeen and save dictionary...')
    tf = TriplesFactory.from_labeled_triples(
        labeled_triples,
        create_inverse_triples=False,
        entity_to_id=None,
        relation_to_id=None,
        compact_id=False,
        filter_out_candidate_inverse_relations=True,
        metadata=None,
    )

    print('Splitting into train/test/validation (seeded, so reproducible)...')
    training, testing, validation = tf.split([.7, .2, .1], random_state=42)

    # Cache the raw labeled triples (so tf can be rebuilt identically) plus
    # the mapped triples needed to reconstruct the split. Plain pickle is
    # used here since this pykeen version (1.6.0) has no to_path_binary /
    # from_path_binary methods on TriplesFactory.
    with open(labeled_triples_path, 'wb') as f:
        pickle.dump(labeled_triples, f)

    split_indices = {
        'train': training.mapped_triples,
        'test': testing.mapped_triples,
        'valid': validation.mapped_triples,
    }
    with open(splits_path, 'wb') as f:
        pickle.dump(split_indices, f)

    print('Cached triples factory and split to disk for future runs.')


# =============================================================================
# Ground truth data (same for both embeddings)
# =============================================================================

print('Loading ground truth CUI pairs...')
gt_data = pd.read_csv(GROUND_TRUTH_PATH, sep=";")
gt_data = gt_data[~gt_data["INTERACTS"].isin([6, 7, 8])]  # keep class 0, exclude unused classes
cuiPairs_full = gt_data["CUI_PAIR"]
pairs_ground_full = gt_data["INTERACTS"]

unique_classes = np.unique(pairs_ground_full)
number_of_classes = len(unique_classes)
print(f"Number of classes: {number_of_classes} -> {unique_classes}")


# =============================================================================
# Excel styling helpers (shared across both embedding runs)
# =============================================================================

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
SUBHEADER_FILL = PatternFill("solid", start_color="2E75B6")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
BOLD_BODY = Font(name="Arial", bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = thin_border


def style_data_row(ws, row, cols, alt=False):
    fill = ALT_FILL if alt else WHITE_FILL
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = BODY_FONT
        cell.alignment = CENTER
        cell.border = thin_border


def write_df_to_sheet(ws, df, start_row=1, title=None):
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = Font(name="Arial", bold=True, size=12, color="1F4E79")
        r += 1
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=r, column=c_idx, value=col_name)
    style_header_row(ws, r, len(df.columns))
    r += 1
    for row_idx, (_, row_data) in enumerate(df.iterrows()):
        alt = row_idx % 2 == 1
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)
        style_data_row(ws, r, len(df.columns), alt=alt)
        r += 1
    return r


def export_results_to_excel(embed_name, df_inner, df_outer, overall_cm, unique_classes,
                             best_sampling, best_hyperparams, best_inner_score_overall,
                             best_features):
    df_summary_rows = []
    for cls in unique_classes:
        df_cls = df_outer[df_outer['Class'] == cls]
        df_summary_rows.append({
            'Class': cls,
            'Avg Precision': round(df_cls['Precision'].mean(), 4),
            'Avg Recall': round(df_cls['Recall'].mean(), 4),
            'Avg F1-Score': round(df_cls['F1-Score'].mean(), 4),
            'Std Precision': round(df_cls['Precision'].std(), 4),
            'Std Recall': round(df_cls['Recall'].std(), 4),
            'Std F1-Score': round(df_cls['F1-Score'].std(), 4),
            'Total Support': int(df_cls['Support'].sum()),
        })
    df_summary = pd.DataFrame(df_summary_rows)

    wb = Workbook()

    # ---- Sheet 1: Summary ----
    ws_summary = wb.active
    ws_summary.title = "Summary"
    r = 1
    ws_summary.cell(row=r, column=1, value=f"{embed_name} -- Nested CV Summary (Outer Folds Average per Class)")
    ws_summary.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_summary.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_summary.columns))
    r += 2
    write_df_to_sheet(ws_summary, df_summary, start_row=r)

    r += len(df_summary) + 3
    ws_summary.cell(row=r, column=1, value="Best Configuration").font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    r += 1
    config = {
        "Best Sampling Config": str(best_sampling),
        "n_estimators": best_hyperparams['n_estimators'],
        "max_depth": best_hyperparams['max_depth'],
        "min_samples_leaf": best_hyperparams['min_samples_leaf'],
        "Best Inner CV F1 (macro)": round(best_inner_score_overall, 4),
        "Num Features Selected": len(best_features) if best_features is not None else "N/A",
    }
    for key, val in config.items():
        ws_summary.cell(row=r, column=1, value=key).font = BOLD_BODY
        ws_summary.cell(row=r, column=1).fill = ALT_FILL
        ws_summary.cell(row=r, column=2, value=val).font = BODY_FONT
        ws_summary.cell(row=r, column=2).fill = WHITE_FILL
        r += 1

    for col in range(1, len(df_summary.columns) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18
    ws_summary.column_dimensions['A'].width = 22

    # ---- Sheet 2: Outer Folds ----
    ws_outer = wb.create_sheet("Outer Folds")
    r = 1
    ws_outer.cell(row=r, column=1, value=f"{embed_name} -- Outer CV Per-Fold Per-Class Metrics")
    ws_outer.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_outer.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_outer.columns))
    r += 2
    write_df_to_sheet(ws_outer, df_outer, start_row=r)
    for col in range(1, len(df_outer.columns) + 1):
        ws_outer.column_dimensions[get_column_letter(col)].width = 20

    # ---- Sheet 3: Inner Folds ----
    ws_inner = wb.create_sheet("Inner Folds")
    r = 1
    ws_inner.cell(row=r, column=1,
                  value=f"{embed_name} -- Inner CV Per-Class Metrics "
                        f"(Outer Fold x Sampling Config x Inner Fold x Class)")
    ws_inner.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_inner.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_inner.columns))
    r += 2
    write_df_to_sheet(ws_inner, df_inner, start_row=r)
    for col in range(1, len(df_inner.columns) + 1):
        ws_inner.column_dimensions[get_column_letter(col)].width = 22

    # ---- Sheet 4: Confusion Matrix ----
    ws_cm = wb.create_sheet("Confusion Matrix")
    ws_cm.cell(row=1, column=1, value=f"{embed_name} -- Overall Confusion Matrix (summed across outer folds)")
    ws_cm.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_cm.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(unique_classes) + 2)

    ws_cm.cell(row=3, column=2, value="Predicted ->").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws_cm.cell(row=3, column=2).fill = HEADER_FILL
    ws_cm.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(unique_classes) + 1)

    ws_cm.cell(row=4, column=1, value="Actual").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws_cm.cell(row=4, column=1).fill = SUBHEADER_FILL
    ws_cm.cell(row=4, column=1).alignment = CENTER

    for c_idx, cls in enumerate(unique_classes, 2):
        cell = ws_cm.cell(row=4, column=c_idx, value=f"Class {cls}")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = thin_border

    ws_cm.cell(row=4, column=len(unique_classes) + 2, value="Support").font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    ws_cm.cell(row=4, column=len(unique_classes) + 2).fill = SUBHEADER_FILL
    ws_cm.cell(row=4, column=len(unique_classes) + 2).alignment = CENTER

    for r_idx, (cls, cm_row) in enumerate(zip(unique_classes, overall_cm)):
        row = 5 + r_idx
        alt = r_idx % 2 == 1
        fill = ALT_FILL if alt else WHITE_FILL

        label_cell = ws_cm.cell(row=row, column=1, value=f"Class {cls}")
        label_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        label_cell.fill = SUBHEADER_FILL
        label_cell.alignment = CENTER

        for c_idx, val in enumerate(cm_row, 2):
            cell = ws_cm.cell(row=row, column=c_idx, value=int(val))
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = CENTER
            cell.border = thin_border

        ws_cm.cell(row=row, column=len(unique_classes) + 2, value=int(np.sum(cm_row))).font = BOLD_BODY
        ws_cm.cell(row=row, column=len(unique_classes) + 2).fill = fill
        ws_cm.cell(row=row, column=len(unique_classes) + 2).alignment = CENTER

    for col in range(1, len(unique_classes) + 3):
        ws_cm.column_dimensions[get_column_letter(col)].width = 14

    output_path = f"{embed_name}_results_corrected_full.xlsx"
    wb.save(output_path)
    print(f"\n[{embed_name}] Results saved to: {output_path}")


# =============================================================================
# Build feature matrix (drug + relation + target embeddings) for a set of pairs
# =============================================================================

def build_feature_matrix(pairs_series, entity_embeddings, rel_embedding, device):
    n = len(pairs_series)
    cols = 3 * len(rel_embedding)
    X = np.zeros((n, cols), dtype=np.float32)
    valid_mask = np.ones(n, dtype=bool)

    for i, pair in enumerate(pairs_series):
        cuiList = pair.split("_")
        drCUI = cuiList[0]
        tarCUI = cuiList[1]

        drId = getIdFromCUI(drCUI)
        if drId is None or drId == "null":
            valid_mask[i] = False
            continue
        tarId = getIdFromCUI(tarCUI)  # FIX: was stale from training loop in original code
        if tarId is None or tarId == "null":
            valid_mask[i] = False
            continue

        dr_entity = torch.as_tensor(tf.entity_to_id[drId], device=device)
        dr_embedding = Variable(entity_embeddings(indices=dr_entity)).cpu().detach().numpy()

        tar_entity = torch.as_tensor(tf.entity_to_id[tarId], device=device)
        tar_embedding = Variable(entity_embeddings(indices=tar_entity)).cpu().detach().numpy()

        X[i] = np.concatenate((dr_embedding, rel_embedding.detach().numpy(), tar_embedding), axis=None)

    return X, valid_mask


# =============================================================================
# Main loop over embeddings
# =============================================================================

for embed in EMBEDDINGS_TO_RUN:
    print(f"\n\n=========================================================")
    print(f" RUNNING EMBEDDING: {embed}")
    print(f"=========================================================\n")

    model_dir = embed
    model_path = os.path.join(model_dir, 'trained_model.pkl')

    if os.path.exists(model_path):
        print(f"[{embed}] Found cached trained model at {model_path}. Loading (skipping training)...")
        model = torch.load(model_path)
    else:
        print(f"[{embed}] No cached model found. Training from scratch...")
        result = pipeline(
            training=training,
            testing=testing,
            validation=validation,
            model=embed,
            stopper='early',
            epochs=100,
            dimensions=100,
            random_seed=42,
            device=DEVICE,
            evaluator_kwargs=dict(batch_size=16),
        )
        result.save_to_directory(model_dir)
        model = result.model
        print(f"[{embed}] Training complete. Saved to {model_dir}/")

    # --- Get relation embedding for INTERACTS_WITH ---
    relID = tf.relation_to_id["INTERACTS_WITH"]
    relation = torch.as_tensor(relID, device=DEVICE)
    relation_representation_modules: List['pykeen.nn.Representation'] = model.relation_representations
    relation_embeddings: pykeen.nn.Embedding = relation_representation_modules[0]
    relation_embedding_tensor = relation_embeddings(indices=relation)
    rel_embedding = Variable(relation_embedding_tensor).cpu()

    entity_representation_modules: List['pykeen.nn.Representation'] = model.entity_representations
    entity_embeddings: pykeen.nn.Embedding = entity_representation_modules[0]

    # --- Build full feature matrix once for all ground-truth pairs ---
    print(f"[{embed}] Building feature matrix for all {len(cuiPairs_full)} ground-truth pairs...")
    X_full, valid_mask = build_feature_matrix(cuiPairs_full, entity_embeddings, rel_embedding, DEVICE)
    save_cui_cache()  # persist CUI->ID cache after building features

    # --- Save the final pairs so that we can run those in BLGPA
    valid_cui_pairs = cuiPairs_full.values[valid_mask]
    pd.DataFrame({'CUI_PAIR':valid_cui_pairs}).to_csv("valid_cui_pairs.csv", index=False)
    print(f"File with {len(valid_cui_pairs)} CUIs saved!!")

    # --- Save the final pairs so that we can run those in BLGPA
    valid_cui_pairs = cuiPairs_full.values[valid_mask]
    pd.DataFrame({'CUI_PAIR':valid_cui_pairs}).to_csv("valid_cui_pairs.csv", index=False)
    print(f"File with {len(valid_cui_pairs)} CUIs saved!!")

    X_all = X_full[valid_mask]
    y_all = pairs_ground_full.values[valid_mask]
    print(f"[{embed}] Valid pairs after dropping unresolved CUIs: {len(y_all)} / {len(cuiPairs_full)}")

    column_names = [f"f{i}" for i in range(X_all.shape[1])]

    # =========================================================================
    # Nested CV (outer 10-fold, inner 5-fold GridSearchCV) with resampling
    # =========================================================================

    outer_cv = StratifiedKFold(n_splits=OUTER_K, shuffle=True, random_state=42)
    inner_cv = StratifiedKFold(n_splits=INNER_K, shuffle=True, random_state=42)

    rf_template = RandomForestClassifier(class_weight="balanced", random_state=42)

    best_inner_score_overall = 0
    best_model = None
    best_sampling = None
    best_hyperparams = None
    best_features = None

    overall_cm = np.zeros((number_of_classes, number_of_classes), dtype=int)
    inner_records = []
    outer_records = []

    for fold_idx, (train_index, test_index) in enumerate(outer_cv.split(X_all, y_all)):
        print(f"\n--- [{embed}] Outer Fold {fold_idx + 1}/{OUTER_K} ---")

        X_train_outer, X_test_outer = X_all[train_index], X_all[test_index]
        y_train_outer, y_test_outer = y_all[train_index], y_all[test_index]

        best_inner_score_fold = 0
        best_inner_model_fold = None
        best_sampling_fold = None
        best_hyperparams_fold = None
        selected_features_mask_fold = None

        X_train_df = pd.DataFrame(X_train_outer, columns=column_names)
        y_train_df = pd.DataFrame(y_train_outer, columns=['INTERACTS'])
        train_data = pd.concat([X_train_df, y_train_df], axis=1)

        for divisor, multiplier in sampling_combos:
            # Class 0 (majority) is downsampled by dividing its count by `divisor`.
            # Classes 1-5 (minority) are SMOTE-upsampled by `multiplier`.
            class_frames = {cls: train_data[train_data['INTERACTS'] == cls] for cls in unique_classes}
            minority_classes = [c for c in unique_classes if c != 0]

            downsample_targets = {
                0: max(1, int(len(class_frames[0]) / divisor))
            }
            upsampled_targets = {
                c: max(1, int(len(class_frames[c]) * multiplier)) for c in minority_classes
            }

            X_for_smote = train_data[column_names]
            y_for_smote = train_data['INTERACTS']

            # Undersample class 0, then SMOTE-upsample classes 1-5, both inside the pipeline
            pipeline_clf = ImbPipeline([
                ("undersample", RandomUnderSampler(sampling_strategy=downsample_targets, random_state=42)),
                ("smote", SMOTE(sampling_strategy=upsampled_targets, random_state=42)),
                ("selector", SelectFromModel(estimator=rf_template, threshold="median")),
                ("clf", RandomForestClassifier(class_weight="balanced", random_state=42))
            ])

            # Pass the RAW, non-resampled outer train data to GridSearchCV
            grid_search_inner = GridSearchCV(
                estimator=pipeline_clf,
                param_grid=param_grid_pipeline,
                cv=inner_cv,
                scoring='f1_macro',
                n_jobs=-1
            )
            grid_search_inner.fit(X_for_smote, y_for_smote)  # <-- RAW data here

            best_params = grid_search_inner.best_params_
            n_estimators = best_params['clf__n_estimators']
            max_depth = best_params['clf__max_depth']
            min_samples_leaf = best_params['clf__min_samples_leaf']

            sampling_label = f"down={divisor},up={multiplier}"

            X_raw_np = X_for_smote.values
            y_raw_np = y_for_smote.values

            # --- Per-inner-fold per-class metrics (5 explicit folds) ---
            for inner_fold_idx, (inner_train_idx, inner_val_idx) in enumerate(
                inner_cv.split(X_for_smote, y_for_smote)
            ):
                X_inner_train = X_raw_np[inner_train_idx]
                y_inner_train = y_raw_np[inner_train_idx]
                X_inner_val = X_raw_np[inner_val_idx]
                y_inner_val = y_raw_np[inner_val_idx]

                inner_fold_pipeline = ImbPipeline([
                    ("undersample", RandomUnderSampler(sampling_strategy=downsample_targets, random_state=42)),
                    ("smote", SMOTE(sampling_strategy=upsampled_targets, random_state=42)),
                    ("selector", SelectFromModel(
                        estimator=RandomForestClassifier(
                            n_estimators=n_estimators, max_depth=max_depth,
                            min_samples_leaf=min_samples_leaf,
                            class_weight="balanced", random_state=42
                        ),
                        threshold="median"
                    )),
                    ("clf", RandomForestClassifier(
                        n_estimators=n_estimators, max_depth=max_depth,
                        min_samples_leaf=min_samples_leaf,
                        class_weight="balanced", random_state=42
                    ))
                ])
                inner_fold_pipeline.fit(X_inner_train, y_inner_train)   # resampling applied only to this fold's train data
                y_pred_inner_val = inner_fold_pipeline.predict(X_inner_val)  # val stays raw/untouched

                inner_report = classification_report(
                    y_inner_val, y_pred_inner_val,
                    labels=unique_classes, output_dict=True, zero_division=0
                )

                for cls in unique_classes:
                    cls_key = str(cls)
                    metrics = inner_report.get(cls_key, {})
                    inner_records.append({
                        'Outer Fold': fold_idx + 1,
                        'Sampling Config': sampling_label,
                        'Inner Fold': inner_fold_idx + 1,
                        'n_estimators': n_estimators,
                        'max_depth': max_depth,
                        'min_samples_leaf': min_samples_leaf,
                        'Class': cls,
                        'Precision': round(metrics.get('precision', 0), 4),
                        'Recall': round(metrics.get('recall', 0), 4),
                        'F1-Score': round(metrics.get('f1-score', 0), 4),
                        'Support': int(metrics.get('support', 0)),
                    })

            # --- Synced pipeline for outer evaluation ---
            synced_pipeline = ImbPipeline([
                ("undersample", RandomUnderSampler(sampling_strategy=downsample_targets, random_state=42)),
                ("smote", SMOTE(sampling_strategy=upsampled_targets, random_state=42)),
                ("selector", SelectFromModel(
                    estimator=RandomForestClassifier(
                        n_estimators=n_estimators, max_depth=max_depth,
                        min_samples_leaf=min_samples_leaf,
                        class_weight="balanced", random_state=42
                    ),
                    threshold="median"
                )),
                ("clf", RandomForestClassifier(
                    n_estimators=n_estimators, max_depth=max_depth,
                    min_samples_leaf=min_samples_leaf,
                    class_weight="balanced", random_state=42
                ))
            ])
            synced_pipeline.fit(X_for_smote, y_for_smote)

            if grid_search_inner.best_score_ > best_inner_score_fold:
                best_inner_score_fold = grid_search_inner.best_score_
                best_inner_model_fold = synced_pipeline
                best_sampling_fold = sampling_label
                best_hyperparams_fold = {
                    "n_estimators": n_estimators,
                    "max_depth": max_depth,
                    "min_samples_leaf": min_samples_leaf
                }
                selector = synced_pipeline.named_steps['selector']
                selected_features_mask_fold = selector.get_support()

        if best_inner_score_fold > best_inner_score_overall:
            best_inner_score_overall = best_inner_score_fold
            best_model = best_inner_model_fold
            best_sampling = best_sampling_fold
            best_hyperparams = best_hyperparams_fold
            best_features = np.array(column_names)[selected_features_mask_fold]

        # --- Outer fold evaluation ---
        y_pred_outer = best_inner_model_fold.predict(X_test_outer)

        outer_score = classification_report(
            y_test_outer, y_pred_outer, labels=unique_classes, output_dict=True, zero_division=0
        )
        cm_outer = confusion_matrix(y_test_outer, y_pred_outer, labels=unique_classes)
        overall_cm += cm_outer

        for cls in unique_classes:
            cls_key = str(cls)
            metrics = outer_score.get(cls_key, {})
            outer_records.append({
                'Outer Fold': fold_idx + 1,
                'Sampling Config': best_sampling_fold,
                'n_estimators': best_hyperparams_fold['n_estimators'],
                'max_depth': best_hyperparams_fold['max_depth'],
                'min_samples_leaf': best_hyperparams_fold['min_samples_leaf'],
                'Class': cls,
                'Precision': round(metrics.get('precision', 0), 4),
                'Recall': round(metrics.get('recall', 0), 4),
                'F1-Score': round(metrics.get('f1-score', 0), 4),
                'Support': int(metrics.get('support', 0)),
            })

    # =========================================================================
    # Export this embedding's results to its own Excel file
    # =========================================================================

    df_inner = pd.DataFrame(inner_records)
    df_outer = pd.DataFrame(outer_records)

    export_results_to_excel(
        embed_name=embed,
        df_inner=df_inner,
        df_outer=df_outer,
        overall_cm=overall_cm,
        unique_classes=unique_classes,
        best_sampling=best_sampling,
        best_hyperparams=best_hyperparams,
        best_inner_score_overall=best_inner_score_overall,
        best_features=best_features,
    )

print("\n\nAll embeddings complete.")
