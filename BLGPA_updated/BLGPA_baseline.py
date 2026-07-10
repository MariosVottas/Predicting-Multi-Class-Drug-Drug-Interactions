import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.datasets import make_classification
from sklearn.model_selection import KFold, cross_val_predict, cross_val_score, StratifiedKFold, GridSearchCV
from sklearn.metrics import confusion_matrix,classification_report, f1_score, accuracy_score
import seaborn as sns
import matplotlib.pyplot as plt
from imblearn.over_sampling import SMOTE
from sklearn.utils import resample
from sklearn.model_selection import train_test_split
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Load the dataset
df1 = pd.read_csv('BLPA_Experiment_Data_8_classes.csv')
df1["INTERACTS"] = df1["INTERACTS"].astype(int)
df1 = df1[~df1["INTERACTS"].isin([6, 7, 8])]

df2 = pd.read_csv('valid_cui_pairs.csv')
# Filter df where drug_pairs matches any drug_pair in df_2
df = df1[df1["CUI_PAIR"].isin(df2["CUI_PAIR"])]

# Keep column names
column_names = df.columns[1:106]

# Separate features from classes
X = df.iloc[:, 1:106].values
y = df.iloc[:, -1].values

# Get the unique class labels
labels = np.unique(y)

# Define a 10-fold cross-validation object
cv = KFold(n_splits=10, shuffle=True, random_state=42)

# --- Collect per-fold, per-class metrics ---
fold_records = []
overall_cm = np.zeros((len(labels), len(labels)), dtype=int)

for fold_idx, (train_idx, test_idx) in enumerate(cv.split(X, y)):
    X_train, X_test = X[train_idx], X[test_idx]
    y_train, y_test = y[train_idx], y[test_idx]

    # Create a random forest classifier with 100 trees
    rfc = RandomForestClassifier(n_estimators=100, random_state=42)
    rfc.fit(X_train, y_train)
    y_pred = rfc.predict(X_test)

    # Per-class report for this fold
    report = classification_report(
        y_test, y_pred, labels=labels, output_dict=True, zero_division=0
    )

    cm = confusion_matrix(y_test, y_pred, labels=labels)
    overall_cm += cm

    for cls in labels:
        cls_key = str(cls)
        metrics = report.get(cls_key, {})
        fold_records.append({
            'Fold': fold_idx + 1,
            'Class': cls,
            'Precision': round(metrics.get('precision', 0), 4),
            'Recall': round(metrics.get('recall', 0), 4),
            'F1-Score': round(metrics.get('f1-score', 0), 4),
            'Support': int(metrics.get('support', 0)),
        })

    print(f"Fold {fold_idx + 1}/{cv.n_splits} done")

df_folds = pd.DataFrame(fold_records)

# --- Summary: average across folds, per class ---
summary_rows = []
for cls in labels:
    df_cls = df_folds[df_folds['Class'] == cls]
    summary_rows.append({
        'Class': cls,
        'Avg Precision': round(df_cls['Precision'].mean(), 4),
        'Avg Recall': round(df_cls['Recall'].mean(), 4),
        'Avg F1-Score': round(df_cls['F1-Score'].mean(), 4),
        'Std Precision': round(df_cls['Precision'].std(), 4),
        'Std Recall': round(df_cls['Recall'].std(), 4),
        'Std F1-Score': round(df_cls['F1-Score'].std(), 4),
        'Total Support': int(df_cls['Support'].sum()),
    })
df_summary = pd.DataFrame(summary_rows)

print("\n=== Average metrics across folds ===")
print(df_summary)


# =============================================================================
# EXCEL EXPORT
# =============================================================================

wb = Workbook()

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
SUBHEADER_FILL = PatternFill("solid", start_color="2E75B6")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
BOLD_BODY = Font(name="Arial", bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = thin_border

def style_data_row(ws, row, cols, alt=False):
    fill = ALT_FILL if alt else WHITE_FILL
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = BODY_FONT
        cell.alignment = CENTER
        cell.border = thin_border

def write_df_to_sheet(ws, df, start_row=1, title=None):
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = Font(name="Arial", bold=True, size=12, color="1F4E79")
        r += 1

    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=r, column=c_idx, value=col_name)
    style_header_row(ws, r, len(df.columns))
    r += 1

    for row_idx, (_, row_data) in enumerate(df.iterrows()):
        alt = row_idx % 2 == 1
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)
        style_data_row(ws, r, len(df.columns), alt=alt)
        r += 1

    return r


# ---- Sheet 1: Summary ----
ws_summary = wb.active
ws_summary.title = "Summary"

r = 1
ws_summary.cell(row=r, column=1, value="10-Fold CV Results — Summary (Average per Class)")
ws_summary.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
ws_summary.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_summary.columns))
r += 2

write_df_to_sheet(ws_summary, df_summary, start_row=r)

for col in range(1, len(df_summary.columns) + 1):
    ws_summary.column_dimensions[get_column_letter(col)].width = 18
ws_summary.column_dimensions['A'].width = 14


# ---- Sheet 2: Per-Fold Results ----
ws_folds = wb.create_sheet("Per-Fold Results")
r = 1
ws_folds.cell(row=r, column=1, value="Per-Fold Per-Class Metrics (10 folds x classes)")
ws_folds.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
ws_folds.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_folds.columns))
r += 2

write_df_to_sheet(ws_folds, df_folds, start_row=r)

for col in range(1, len(df_folds.columns) + 1):
    ws_folds.column_dimensions[get_column_letter(col)].width = 16


# ---- Sheet 3: Confusion Matrix ----
ws_cm = wb.create_sheet("Confusion Matrix")
ws_cm.cell(row=1, column=1, value="Overall Confusion Matrix (summed across all 10 folds)")
ws_cm.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
ws_cm.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(labels) + 2)

ws_cm.cell(row=3, column=2, value="Predicted ->").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws_cm.cell(row=3, column=2).fill = HEADER_FILL
ws_cm.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(labels) + 1)

ws_cm.cell(row=4, column=1, value="Actual").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws_cm.cell(row=4, column=1).fill = SUBHEADER_FILL
ws_cm.cell(row=4, column=1).alignment = CENTER

for c_idx, cls in enumerate(labels, 2):
    cell = ws_cm.cell(row=4, column=c_idx, value=f"Class {cls}")
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = thin_border

ws_cm.cell(row=4, column=len(labels) + 2, value="Support").font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ws_cm.cell(row=4, column=len(labels) + 2).fill = SUBHEADER_FILL
ws_cm.cell(row=4, column=len(labels) + 2).alignment = CENTER

for r_idx, (cls, cm_row) in enumerate(zip(labels, overall_cm)):
    row = 5 + r_idx
    alt = r_idx % 2 == 1
    fill = ALT_FILL if alt else WHITE_FILL

    label_cell = ws_cm.cell(row=row, column=1, value=f"Class {cls}")
    label_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    label_cell.fill = SUBHEADER_FILL
    label_cell.alignment = CENTER

    for c_idx, val in enumerate(cm_row, 2):
        cell = ws_cm.cell(row=row, column=c_idx, value=int(val))
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = thin_border

    ws_cm.cell(row=row, column=len(labels) + 2, value=int(np.sum(cm_row))).font = BOLD_BODY
    ws_cm.cell(row=row, column=len(labels) + 2).fill = fill
    ws_cm.cell(row=row, column=len(labels) + 2).alignment = CENTER

for col in range(1, len(labels) + 3):
    ws_cm.column_dimensions[get_column_letter(col)].width = 14

# output_path = "simple_cv_results.xlsx"
output_path = "results_baseline.xlsx"
wb.save(output_path)
print(f"\nResults saved to: {output_path}")