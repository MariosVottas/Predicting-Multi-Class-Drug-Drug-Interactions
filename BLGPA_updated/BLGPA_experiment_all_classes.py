import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.datasets import make_classification
from sklearn.model_selection import KFold, cross_val_predict, cross_val_score, StratifiedKFold, GridSearchCV
from sklearn.metrics import confusion_matrix,classification_report, f1_score, accuracy_score
import seaborn as sns
import matplotlib.pyplot as plt
from imblearn.over_sampling import SMOTE
from sklearn.utils import resample
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import SelectFromModel
from sklearn.pipeline import Pipeline
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Load the dataset
df1 = pd.read_csv('BLPA_Experiment_Data_8_classes.csv')
df1["INTERACTS"] = df1["INTERACTS"].astype(int)
df1 = df1[~df1["INTERACTS"].isin([6, 7, 8])]

df2 = pd.read_csv('valid_cui_pairs.csv')
# Filter df where drug_pairs matches any drug_pair in df_2
df = df1[df1["CUI_PAIR"].isin(df2["CUI_PAIR"])]

# df['INTERACTS'].value_counts().reset_index().to_csv('class_counts.csv', index=False)
# print("Support saved!")
# Keep column names
column_names = df.columns[1:106]

# Separate features from classes
X = df.iloc[:, 1:106].values
y = df.iloc[:, -1].values

# X, X_test, y, y_test = train_test_split(X1, y1, test_size=0.2, stratify = y1 )


# --- ASSUMPTIONS (Ensure these variables are defined in your environment) ---
# X: Your feature matrix (e.g., NumPy array)
# y: Your target vector (e.g., NumPy array)
# column_names: List of column names corresponding to X
# ----------------------------------------------------------------------------

# --- Model & hyperparameter grid ---
rf_template = RandomForestClassifier(class_weight="balanced", random_state=42)

param_grid_pipeline = {
    'clf__n_estimators': [100, 200, 500],
    'clf__max_depth': [5, 10],
    'clf__min_samples_leaf': [5, 10],
}

sampling_values = [[5, 2], [10, 2], [5,3], [10,3]]

# --- Cross-validation setup ---
outer_cv = StratifiedKFold(n_splits=10, shuffle=True, random_state=42)
inner_cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

# --- Tracking variables ---
unique_classes = np.unique(y)
outer_scores_list = []

best_inner_score_overall = 0
best_model = None
best_sampling = None
best_hyperparams = None
best_features = None

overall_cm = np.zeros((len(unique_classes), len(unique_classes)), dtype=int)
overall_supports = np.zeros(len(unique_classes), dtype=int)
cumulative_recalls = {class_label: 0 for class_label in unique_classes}
cumulative_precisions = {class_label: 0 for class_label in unique_classes}
cumulative_f1_scores = {class_label: 0 for class_label in unique_classes}
class_counts = {class_label: 0 for class_label in unique_classes}

# --- NEW: collectors for Excel export ---
# Inner fold records: one row per (outer_fold, sampling_config, inner_fold, class)
inner_records = []
# Outer fold records: one row per (outer_fold, class)
outer_records = []


# --- Nested CV ---
for fold_idx, (train_index, test_index) in enumerate(outer_cv.split(X, y)):
    print(f"\n--- Starting Outer Fold {fold_idx + 1}/{outer_cv.n_splits} ---")

    X_train_outer, X_test_outer = X[train_index], X[test_index]
    y_train_outer, y_test_outer = y[train_index], y[test_index]

    best_inner_score_fold = 0
    best_inner_model_fold = None

    X_train_df = pd.DataFrame(X_train_outer, columns=column_names)
    y_train_df = pd.DataFrame(y_train_outer, columns=['INTERACTS'])
    train_data = pd.concat([X_train_df, y_train_df], axis=1)

    for i, j in sampling_values:
        majority_class_0 = train_data[train_data['INTERACTS'] == 0]
        minority_class_1 = train_data[train_data['INTERACTS'] == 1]
        minority_class_2 = train_data[train_data['INTERACTS'] == 2]
        minority_class_3 = train_data[train_data['INTERACTS'] == 3]
        minority_class_4 = train_data[train_data['INTERACTS'] == 4]
        minority_class_5 = train_data[train_data['INTERACTS'] == 5]

        majority_downsampled_samples = len(majority_class_0) // i
        minority_upsampled_samples = {
            1: int(len(minority_class_1) * j),
            2: int(len(minority_class_2) * j),
            3: int(len(minority_class_3) * j),
            4: int(len(minority_class_4) * j),
            5: int(len(minority_class_5) * j),
        }

        majority_downsampled = resample(
            majority_class_0, replace=False,
            n_samples=majority_downsampled_samples, random_state=42
        )

        minority_combined = pd.concat([
            minority_class_1, minority_class_2,
            minority_class_3, minority_class_4, minority_class_5
        ])

        smote = SMOTE(sampling_strategy=minority_upsampled_samples, random_state=42)
        X_minority = minority_combined[column_names]
        y_minority = minority_combined['INTERACTS']
        X_minority_upsampled, y_minority_upsampled = smote.fit_resample(X_minority, y_minority)

        minority_upsampled = pd.DataFrame(X_minority_upsampled, columns=column_names)
        minority_upsampled['INTERACTS'] = y_minority_upsampled

        df_downsampled = pd.concat(
            [majority_downsampled, minority_upsampled], ignore_index=True
        ).sample(frac=1, random_state=42)

        X_train_upsampled = df_downsampled[column_names].values
        y_train_upsampled = df_downsampled['INTERACTS'].values

        pipeline = Pipeline([
            ("selector", SelectFromModel(estimator=rf_template, threshold="median")),
            ("clf", RandomForestClassifier(class_weight="balanced", random_state=42))
        ])

        grid_search_inner = GridSearchCV(
            estimator=pipeline,
            param_grid=param_grid_pipeline,
            cv=inner_cv,
            scoring='f1_macro',
            refit=True,
            n_jobs=-1,
            return_train_score=False
        )

        grid_search_inner.fit(X_train_upsampled, y_train_upsampled)

        best_params = grid_search_inner.best_params_
        n_estimators = best_params['clf__n_estimators']
        max_depth = best_params['clf__max_depth']
        min_samples_leaf = best_params['clf__min_samples_leaf']

        synced_pipeline = Pipeline([
            ("selector", SelectFromModel(
                estimator=RandomForestClassifier(
                    n_estimators=n_estimators, max_depth=max_depth,
                    min_samples_leaf=min_samples_leaf,
                    class_weight="balanced", random_state=42
                ),
                threshold="median"
            )),
            ("clf", RandomForestClassifier(
                n_estimators=n_estimators, max_depth=max_depth,
                min_samples_leaf=min_samples_leaf,
                class_weight="balanced", random_state=42
            ))
        ])
        synced_pipeline.fit(X_train_upsampled, y_train_upsampled)

        # --- NEW: Collect per-class metrics for each of the 5 inner folds ---
        # Re-run the best synced pipeline on each inner split explicitly
        sampling_label = f"down={i}_up={j}"
        for inner_fold_idx, (inner_train_idx, inner_val_idx) in enumerate(
            inner_cv.split(X_train_upsampled, y_train_upsampled)
        ):
            X_inner_train = X_train_upsampled[inner_train_idx]
            y_inner_train = y_train_upsampled[inner_train_idx]
            X_inner_val   = X_train_upsampled[inner_val_idx]
            y_inner_val   = y_train_upsampled[inner_val_idx]

            # Build pipeline with the best hyperparams found by GridSearchCV
            inner_fold_pipeline = Pipeline([
                ("selector", SelectFromModel(
                    estimator=RandomForestClassifier(
                        n_estimators=n_estimators, max_depth=max_depth,
                        min_samples_leaf=min_samples_leaf,
                        class_weight="balanced", random_state=42
                    ),
                    threshold="median"
                )),
                ("clf", RandomForestClassifier(
                    n_estimators=n_estimators, max_depth=max_depth,
                    min_samples_leaf=min_samples_leaf,
                    class_weight="balanced", random_state=42
                ))
            ])
            inner_fold_pipeline.fit(X_inner_train, y_inner_train)
            y_pred_inner_val = inner_fold_pipeline.predict(X_inner_val)

            inner_report = classification_report(
                y_inner_val, y_pred_inner_val,
                labels=unique_classes, output_dict=True, zero_division=0
            )

            for cls in unique_classes:
                cls_key = str(cls)
                metrics = inner_report.get(cls_key, {})
                inner_records.append({
                    'Outer Fold': fold_idx + 1,
                    'Sampling Config': sampling_label,
                    'Inner Fold': inner_fold_idx + 1,
                    'n_estimators': n_estimators,
                    'max_depth': max_depth,
                    'min_samples_leaf': min_samples_leaf,
                    'Class': cls,
                    'Precision': round(metrics.get('precision', 0), 4),
                    'Recall': round(metrics.get('recall', 0), 4),
                    'F1-Score': round(metrics.get('f1-score', 0), 4),
                    'Support': int(metrics.get('support', 0)),
                })

        if grid_search_inner.best_score_ > best_inner_score_fold:
            best_inner_score_fold = grid_search_inner.best_score_
            best_inner_model_fold = synced_pipeline
            best_sampling_fold = (i, j)
            best_hyperparams_fold = {
                "n_estimators": n_estimators,
                "max_depth": max_depth,
                "min_samples_leaf": min_samples_leaf
            }
            selector = best_inner_model_fold.named_steps['selector']
            selected_features_mask_fold = selector.get_support()

    if best_inner_score_fold > best_inner_score_overall:
        best_inner_score_overall = best_inner_score_fold
        best_model = best_inner_model_fold
        best_sampling = best_sampling_fold
        best_hyperparams = best_hyperparams_fold
        best_features = np.array(column_names)[selected_features_mask_fold]

    # --- Outer fold evaluation ---
    y_pred_outer = best_inner_model_fold.predict(X_test_outer)

    outer_score = classification_report(
        y_test_outer, y_pred_outer, labels=unique_classes, output_dict=True
    )
    cm_outer = confusion_matrix(y_test_outer, y_pred_outer, labels=unique_classes)

    overall_cm += cm_outer
    overall_supports += np.sum(cm_outer, axis=1)
    outer_scores_list.append(outer_score)

    # --- NEW: Collect per-class metrics from outer fold ---
    for cls in unique_classes:
        cls_key = str(cls)
        metrics = outer_score.get(cls_key, {})
        outer_records.append({
            'Outer Fold': fold_idx + 1,
            'Sampling Config': f"down={best_sampling_fold[0]}_up={best_sampling_fold[1]}",
            'n_estimators': best_hyperparams_fold['n_estimators'],
            'max_depth': best_hyperparams_fold['max_depth'],
            'min_samples_leaf': best_hyperparams_fold['min_samples_leaf'],
            'Class': cls,
            'Precision': round(metrics.get('precision', 0), 4),
            'Recall': round(metrics.get('recall', 0), 4),
            'F1-Score': round(metrics.get('f1-score', 0), 4),
            'Support': int(metrics.get('support', 0)),
        })

    for class_label in unique_classes:
        class_key = str(class_label)
        class_metrics = outer_score.get(class_key, {})
        cumulative_recalls[class_label] += class_metrics.get('recall', 0)
        cumulative_precisions[class_label] += class_metrics.get('precision', 0)
        cumulative_f1_scores[class_label] += class_metrics.get('f1-score', 0)
        class_counts[class_label] += 1

# --- Final Summary (console) ---
print("\n===========================")
print(" OVERALL RESULTS SUMMARY ")
print("===========================\n")
print("Overall Confusion Matrix:")
print(overall_cm)

average_recalls = {cls: cumulative_recalls[cls] / count for cls, count in class_counts.items() if count > 0}
average_precisions = {cls: cumulative_precisions[cls] / count for cls, count in class_counts.items() if count > 0}
average_f1_scores = {cls: cumulative_f1_scores[cls] / count for cls, count in class_counts.items() if count > 0}

for class_label in unique_classes:
    if class_label in average_f1_scores:
        print(f"\nClass '{class_label}':")
        print(f"Average F1 Score: {average_f1_scores[class_label]:.4f}")
        print(f"Average Recall: {average_recalls[class_label]:.4f}")
        print(f"Average Precision: {average_precisions[class_label]:.4f}")

print("\n===========================")
print(" BEST CONFIGURATION SUMMARY ")
print("===========================\n")
print(f"Best Upsampling/Downsampling Ratio: {best_sampling}")
print(f"Best Hyperparameters: {best_hyperparams}")
print(f"Best Inner CV F1 (macro): {best_inner_score_overall:.4f}")
print(f"Number of Features Selected: {len(best_features) if best_features is not None else 'N/A'}")
print("Selected Features (from the best inner model):")
print(best_features)


# =============================================================================
# EXCEL EXPORT
# =============================================================================

df_inner = pd.DataFrame(inner_records)
df_outer = pd.DataFrame(outer_records)

# --- Summary: average outer metrics per class ---
summary_rows = []
for cls in unique_classes:
    df_cls = df_outer[df_outer['Class'] == cls]
    summary_rows.append({
        'Class': cls,
        'Avg Precision': round(df_cls['Precision'].mean(), 4),
        'Avg Recall': round(df_cls['Recall'].mean(), 4),
        'Avg F1-Score': round(df_cls['F1-Score'].mean(), 4),
        'Std Precision': round(df_cls['Precision'].std(), 4),
        'Std Recall': round(df_cls['Recall'].std(), 4),
        'Std F1-Score': round(df_cls['F1-Score'].std(), 4),
        'Total Support': int(df_cls['Support'].sum()),
    })
df_summary = pd.DataFrame(summary_rows)

# --- Build workbook ---
wb = Workbook()

# Styling helpers
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
SUBHEADER_FILL = PatternFill("solid", start_color="2E75B6")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_BODY = Font(name="Arial", bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = thin_border

def style_data_row(ws, row, cols, alt=False):
    fill = ALT_FILL if alt else WHITE_FILL
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = BODY_FONT
        cell.alignment = CENTER
        cell.border = thin_border

def write_df_to_sheet(ws, df, start_row=1, title=None):
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = Font(name="Arial", bold=True, size=12, color="1F4E79")
        r += 1

    # Header
    for c_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=r, column=c_idx, value=col_name)
    style_header_row(ws, r, len(df.columns))
    r += 1

    # Data
    for row_idx, (_, row_data) in enumerate(df.iterrows()):
        alt = row_idx % 2 == 1
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)
        style_data_row(ws, r, len(df.columns), alt=alt)
        r += 1

    return r


# ---- Sheet 1: Summary ----
ws_summary = wb.active
ws_summary.title = "Summary"

r = 1
ws_summary.cell(row=r, column=1, value="Nested CV Results — Summary (Outer Folds Average per Class)")
ws_summary.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
ws_summary.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_summary.columns))
r += 2

write_df_to_sheet(ws_summary, df_summary, start_row=r)

# Best config block
r += len(df_summary) + 3
ws_summary.cell(row=r, column=1, value="Best Configuration").font = Font(name="Arial", bold=True, size=12, color="1F4E79")
r += 1
config = {
    "Best Sampling (down, up)": str(best_sampling),
    "n_estimators": best_hyperparams['n_estimators'],
    "max_depth": best_hyperparams['max_depth'],
    "min_samples_leaf": best_hyperparams['min_samples_leaf'],
    "Best Inner CV F1 (macro)": round(best_inner_score_overall, 4),
    "Num Features Selected": len(best_features) if best_features is not None else "N/A",
    "Selected Features": ", ".join(best_features) if best_features is not None else "N/A",
}
for key, val in config.items():
    ws_summary.cell(row=r, column=1, value=key).font = BOLD_BODY
    ws_summary.cell(row=r, column=1).fill = ALT_FILL
    ws_summary.cell(row=r, column=2, value=val).font = BODY_FONT
    ws_summary.cell(row=r, column=2).fill = WHITE_FILL
    r += 1

# Set column widths
for col in range(1, len(df_summary.columns) + 1):
    ws_summary.column_dimensions[get_column_letter(col)].width = 18
ws_summary.column_dimensions['A'].width = 22


# ---- Sheet 2: Outer Folds (per fold × class) ----
ws_outer = wb.create_sheet("Outer Folds")
r = 1
ws_outer.cell(row=r, column=1, value="Outer CV — Per-Fold Per-Class Metrics")
ws_outer.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
ws_outer.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_outer.columns))
r += 2

write_df_to_sheet(ws_outer, df_outer, start_row=r)

for col in range(1, len(df_outer.columns) + 1):
    ws_outer.column_dimensions[get_column_letter(col)].width = 20
ws_outer.column_dimensions['A'].width = 12


# ---- Sheet 3: Inner Folds (per sampling × outer fold × class) ----
ws_inner = wb.create_sheet("Inner Folds")
r = 1
ws_inner.cell(row=r, column=1, value="Inner CV — Per-Class Metrics per Outer Fold x Sampling Config x Inner Fold (5 folds x 2 configs x 6 classes = 60 rows per outer fold)")
ws_inner.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
ws_inner.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_inner.columns))
r += 2

write_df_to_sheet(ws_inner, df_inner, start_row=r)

for col in range(1, len(df_inner.columns) + 1):
    ws_inner.column_dimensions[get_column_letter(col)].width = 22
ws_inner.column_dimensions['A'].width = 12


# ---- Sheet 4: Confusion Matrix ----
ws_cm = wb.create_sheet("Confusion Matrix")
ws_cm.cell(row=1, column=1, value="Overall Confusion Matrix (sum across outer folds)")
ws_cm.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
ws_cm.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(unique_classes) + 2)

ws_cm.cell(row=3, column=1, value="").font = BOLD_BODY
ws_cm.cell(row=3, column=2, value="Predicted →").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws_cm.cell(row=3, column=2).fill = HEADER_FILL
ws_cm.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(unique_classes) + 1)

ws_cm.cell(row=4, column=1, value="Actual ↓").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ws_cm.cell(row=4, column=1).fill = SUBHEADER_FILL
ws_cm.cell(row=4, column=1).alignment = CENTER

for c_idx, cls in enumerate(unique_classes, 2):
    cell = ws_cm.cell(row=4, column=c_idx, value=f"Class {cls}")
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = thin_border

# Support column header
ws_cm.cell(row=4, column=len(unique_classes) + 2, value="Support").font = SUBHEADER_FONT
ws_cm.cell(row=4, column=len(unique_classes) + 2).fill = SUBHEADER_FILL
ws_cm.cell(row=4, column=len(unique_classes) + 2).alignment = CENTER

for r_idx, (cls, cm_row) in enumerate(zip(unique_classes, overall_cm)):
    row = 5 + r_idx
    alt = r_idx % 2 == 1
    fill = ALT_FILL if alt else WHITE_FILL

    label_cell = ws_cm.cell(row=row, column=1, value=f"Class {cls}")
    label_cell.font = BOLD_BODY
    label_cell.fill = SUBHEADER_FILL
    label_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    label_cell.alignment = CENTER

    for c_idx, val in enumerate(cm_row, 2):
        cell = ws_cm.cell(row=row, column=c_idx, value=int(val))
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = thin_border

    ws_cm.cell(row=row, column=len(unique_classes) + 2, value=int(np.sum(cm_row))).font = BOLD_BODY
    ws_cm.cell(row=row, column=len(unique_classes) + 2).fill = fill
    ws_cm.cell(row=row, column=len(unique_classes) + 2).alignment = CENTER

for col in range(1, len(unique_classes) + 3):
    ws_cm.column_dimensions[get_column_letter(col)].width = 14


output_path = "results_BLGPA.xlsx"
wb.save(output_path)
print(f"\nResults saved to: {output_path}")
