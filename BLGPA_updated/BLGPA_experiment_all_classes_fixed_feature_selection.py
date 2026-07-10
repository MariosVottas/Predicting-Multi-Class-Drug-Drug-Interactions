import pandas as pd
import numpy as np
from collections import Counter

from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import StratifiedKFold, GridSearchCV, train_test_split
from sklearn.metrics import confusion_matrix, classification_report
from sklearn.feature_selection import SelectFromModel, RFE, SelectKBest, f_classif
from sklearn.base import clone

from imblearn.pipeline import Pipeline as ImbPipeline
from imblearn.over_sampling import SMOTE
from imblearn.under_sampling import RandomUnderSampler

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# SAFE SMOTE
# =============================================================================
# SMOTE now runs on much smaller slices of data than before (only the
# inner-CV *training* fold, instead of the whole outer-training fold), which
# makes it more likely that some minority class has fewer rows than the
# default k_neighbors=5 requires. This subclass caps k_neighbors at fit time
# so a rare class in a small fold can't crash the run.
class SafeSMOTE(SMOTE):
    def _fit_resample(self, X, y):
        y_counts = Counter(y)
        target_classes = list(self.sampling_strategy_.keys())
        relevant_counts = [y_counts[c] for c in target_classes if y_counts.get(c, 0) > 0]
        if relevant_counts:
            base_k = self.k_neighbors if isinstance(self.k_neighbors, int) else 5
            safe_k = max(1, min(relevant_counts) - 1)
            self.k_neighbors = min(base_k, safe_k)
        return super()._fit_resample(X, y)


def make_downsample_strategy(ratio):
    """Callable sampling_strategy for RandomUnderSampler: shrink class 0
    to (count // ratio), same rule as the original script, but evaluated
    fresh on whatever y is passed in (i.e. the current CV training fold)."""
    def strategy(y):
        counts = Counter(y)
        n0 = counts.get(0, 0)
        if n0 == 0:
            return {}
        return {0: max(1, n0 // ratio)}
    return strategy


def make_upsample_strategy(ratio):
    """Callable sampling_strategy for SMOTE: grow each minority class
    (1-5) to (count * ratio), same rule as the original script, evaluated
    fresh on the current CV training fold."""
    def strategy(y):
        counts = Counter(y)
        target = {}
        for cls in [1, 2, 3, 4, 5]:
            if counts.get(cls, 0) > 0:
                target[cls] = int(counts[cls] * ratio)
        return target
    return strategy


# =============================================================================
# DATA LOADING & PREP
# =============================================================================
df1 = pd.read_csv('BLPA_Experiment_Data_8_classes.csv')
df1["INTERACTS"] = df1["INTERACTS"].astype(int)
df1 = df1[~df1["INTERACTS"].isin([6, 7, 8])]

df2 = pd.read_csv('valid_cui_pairs.csv')
df = df1[df1["CUI_PAIR"].isin(df2["CUI_PAIR"])]

column_names = df.columns[1:106]
X1 = df.iloc[:, 1:106].values
y1 = df.iloc[:, -1].values

X, X_test, y, y_test = train_test_split(X1, y1, test_size=0.2, stratify=y1, random_state=42)
# NOTE: X_test / y_test are a genuine untouched holdout, same as your
# original script. They are still not used anywhere below -- left as-is
# since that matches your original structure. Let me know if you'd like
# a final "evaluate best overall model on this holdout" step added.

# =============================================================================
# EXPERIMENT CONFIGURATION
# =============================================================================
rf_template = RandomForestClassifier(class_weight="balanced", random_state=42)

# param_grid_pipeline = {
#     'clf__n_estimators': [100, 200, 500],
#     'clf__max_depth': [5, 10],
#     'clf__min_samples_leaf': [5, 10],
# }
param_grid_pipeline = {
    'clf__n_estimators': [500],
    'clf__max_depth': [10],
    'clf__min_samples_leaf': [5],
}

# (downsample_ratio, upsample_ratio) pairs, same grid as the original script
# sampling_values = [[5, 2], [10, 2], [5, 3], [10, 3]]
sampling_values = [[10, 3]]

# The 3 feature selection methods to cycle through (same as file 2)
feature_selection_methods = {
    # "SelectFromModel": {
    #     "selector": SelectFromModel(estimator=rf_template, threshold="median"),
    #     "extra_grid": {}
    # },
    "RFE": {
        "selector": RFE(estimator=rf_template, step=5),
        "extra_grid": {'selector__n_features_to_select': [20, 50, 80]}
    } # ,
    # "ANOVA_Filter": {
    #     "selector": SelectKBest(score_func=f_classif),
    #     "extra_grid": {'selector__k': [20, 50, 80]}
    # }
}

outer_cv = StratifiedKFold(n_splits=10, shuffle=True, random_state=42)
inner_cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

# =============================================================================
# MAIN EXPERIMENT LOOP (one full nested CV run per feature selection method)
# =============================================================================
for method_name, method_config in feature_selection_methods.items():
    print(f"\n{'='*60}")
    print(f" STARTING EXPERIMENT: {method_name} ")
    print(f"{'='*60}\n")

    unique_classes = np.unique(y)
    outer_scores_list = []

    best_inner_score_overall = 0
    best_model = None
    best_sampling = None
    best_hyperparams = None
    best_features = None

    overall_cm = np.zeros((len(unique_classes), len(unique_classes)), dtype=int)
    overall_supports = np.zeros(len(unique_classes), dtype=int)
    cumulative_recalls = {cls: 0 for cls in unique_classes}
    cumulative_precisions = {cls: 0 for cls in unique_classes}
    cumulative_f1_scores = {cls: 0 for cls in unique_classes}
    class_counts = {cls: 0 for cls in unique_classes}

    inner_records = []
    outer_records = []

    current_param_grid = {**param_grid_pipeline, **method_config["extra_grid"]}

    # --- Nested CV ---
    for fold_idx, (train_index, test_index) in enumerate(outer_cv.split(X, y)):
        print(f"--- Starting Outer Fold {fold_idx + 1}/{outer_cv.n_splits} ---")

        X_train_outer, X_test_outer = X[train_index], X[test_index]
        y_train_outer, y_test_outer = y[train_index], y[test_index]

        best_inner_score_fold = 0
        best_inner_model_fold = None
        best_sampling_fold = None
        best_hyperparams_fold = None
        selected_features_mask_fold = None

        for i, j in sampling_values:
            sampling_label = f"down={i}_up={j}"

            # Resampling now lives INSIDE the pipeline. imblearn's Pipeline
            # only calls fit_resample() during .fit() -- during .predict()/
            # .score() the resampling steps are skipped and the real,
            # untouched data passes straight to the selector/classifier.
            # That means GridSearchCV's inner CV (and the manual inner-fold
            # loop below) each resample only their own training fold, never
            # touching the corresponding validation fold. No more leakage.
            pipeline = ImbPipeline([
                ("undersample", RandomUnderSampler(
                    sampling_strategy=make_downsample_strategy(i), random_state=42
                )),
                ("oversample", SafeSMOTE(
                    sampling_strategy=make_upsample_strategy(j), random_state=42
                )),
                ("selector", clone(method_config["selector"])),
                ("clf", RandomForestClassifier(class_weight="balanced", random_state=42))
            ])

            grid_search_inner = GridSearchCV(
                estimator=pipeline,
                param_grid=current_param_grid,
                cv=inner_cv,
                scoring='f1_macro',
                refit=True,
                n_jobs=1,
                return_train_score=False
            )

            # Fit on the RAW outer-training fold -- resampling happens
            # internally, per inner-CV split, inside the pipeline.
            grid_search_inner.fit(X_train_outer, y_train_outer)

            best_inner_model_fold = grid_search_inner.best_estimator_
            rf_step = best_inner_model_fold.named_steps['clf']
            n_estimators = rf_step.n_estimators
            max_depth = rf_step.max_depth
            min_samples_leaf = rf_step.min_samples_leaf

            # --- Explicit inner-fold metrics, still leak-free ---
            # Split the RAW outer-training fold (never the resampled data)
            # and let each cloned pipeline do its own resampling on just
            # its training slice.
            for inner_fold_idx, (inner_train_idx, inner_val_idx) in enumerate(
                inner_cv.split(X_train_outer, y_train_outer)
            ):
                X_inner_train = X_train_outer[inner_train_idx]
                y_inner_train = y_train_outer[inner_train_idx]
                X_inner_val = X_train_outer[inner_val_idx]
                y_inner_val = y_train_outer[inner_val_idx]

                inner_fold_pipeline = clone(best_inner_model_fold)
                inner_fold_pipeline.fit(X_inner_train, y_inner_train)
                y_pred_inner_val = inner_fold_pipeline.predict(X_inner_val)

                inner_report = classification_report(
                    y_inner_val, y_pred_inner_val,
                    labels=unique_classes, output_dict=True, zero_division=0
                )

                for cls in unique_classes:
                    cls_key = str(cls)
                    metrics = inner_report.get(cls_key, {})
                    inner_records.append({
                        'Outer Fold': fold_idx + 1,
                        'Sampling Config': sampling_label,
                        'Inner Fold': inner_fold_idx + 1,
                        'n_estimators': n_estimators,
                        'max_depth': max_depth,
                        'min_samples_leaf': min_samples_leaf,
                        'Class': cls,
                        'Precision': round(metrics.get('precision', 0), 4),
                        'Recall': round(metrics.get('recall', 0), 4),
                        'F1-Score': round(metrics.get('f1-score', 0), 4),
                        'Support': int(metrics.get('support', 0)),
                    })

            if grid_search_inner.best_score_ > best_inner_score_fold:
                best_inner_score_fold = grid_search_inner.best_score_
                best_sampling_fold = (i, j)
                best_hyperparams_fold = {
                    "n_estimators": n_estimators,
                    "max_depth": max_depth,
                    "min_samples_leaf": min_samples_leaf
                }
                selector_step = best_inner_model_fold.named_steps['selector']
                selected_features_mask_fold = selector_step.get_support()

        if best_inner_score_fold > best_inner_score_overall:
            best_inner_score_overall = best_inner_score_fold
            best_model = best_inner_model_fold
            best_sampling = best_sampling_fold
            best_hyperparams = best_hyperparams_fold
            best_features = np.array(column_names)[selected_features_mask_fold]

        # --- Outer fold evaluation (untouched real data in, untouched real data out) ---
        y_pred_outer = best_inner_model_fold.predict(X_test_outer)

        outer_score = classification_report(
            y_test_outer, y_pred_outer, labels=unique_classes, output_dict=True, zero_division=0
        )
        cm_outer = confusion_matrix(y_test_outer, y_pred_outer, labels=unique_classes)

        overall_cm += cm_outer
        overall_supports += np.sum(cm_outer, axis=1)
        outer_scores_list.append(outer_score)

        for cls in unique_classes:
            cls_key = str(cls)
            metrics = outer_score.get(cls_key, {})
            outer_records.append({
                'Outer Fold': fold_idx + 1,
                'Sampling Config': f"down={best_sampling_fold[0]}_up={best_sampling_fold[1]}",
                'n_estimators': best_hyperparams_fold['n_estimators'],
                'max_depth': best_hyperparams_fold['max_depth'],
                'min_samples_leaf': best_hyperparams_fold['min_samples_leaf'],
                'Class': cls,
                'Precision': round(metrics.get('precision', 0), 4),
                'Recall': round(metrics.get('recall', 0), 4),
                'F1-Score': round(metrics.get('f1-score', 0), 4),
                'Support': int(metrics.get('support', 0)),
            })

        for class_label in unique_classes:
            class_key = str(class_label)
            class_metrics = outer_score.get(class_key, {})
            cumulative_recalls[class_label] += class_metrics.get('recall', 0)
            cumulative_precisions[class_label] += class_metrics.get('precision', 0)
            cumulative_f1_scores[class_label] += class_metrics.get('f1-score', 0)
            class_counts[class_label] += 1

    # =============================================================================
    # CONSOLE SUMMARY
    # =============================================================================
    print(f"\n[{method_name}] OVERALL RESULTS SUMMARY")
    print("=======================================")
    print(overall_cm)

    average_recalls = {cls: cumulative_recalls[cls] / count for cls, count in class_counts.items() if count > 0}
    average_precisions = {cls: cumulative_precisions[cls] / count for cls, count in class_counts.items() if count > 0}
    average_f1_scores = {cls: cumulative_f1_scores[cls] / count for cls, count in class_counts.items() if count > 0}

    for class_label in unique_classes:
        if class_label in average_f1_scores:
            print(f"\nClass '{class_label}':")
            print(f"Average F1 Score: {average_f1_scores[class_label]:.4f}")
            print(f"Average Recall: {average_recalls[class_label]:.4f}")
            print(f"Average Precision: {average_precisions[class_label]:.4f}")

    print(f"\n[{method_name}] BEST CONFIGURATION SUMMARY")
    print("=======================================")
    print(f"Best Downsampling/Upsampling Ratio: {best_sampling}")
    print(f"Best Hyperparameters: {best_hyperparams}")
    print(f"Best Inner CV F1 (macro): {best_inner_score_overall:.4f}")
    print(f"Number of Features Selected: {len(best_features) if best_features is not None else 'N/A'}")
    print("Selected Features (from the best inner model):")
    print(best_features)

    # =============================================================================
    # EXCEL EXPORT (one workbook per feature selection method)
    # =============================================================================
    df_inner = pd.DataFrame(inner_records)
    df_outer = pd.DataFrame(outer_records)

    summary_rows = []
    for cls in unique_classes:
        df_cls = df_outer[df_outer['Class'] == cls]
        summary_rows.append({
            'Class': cls,
            'Avg Precision': round(df_cls['Precision'].mean(), 4),
            'Avg Recall': round(df_cls['Recall'].mean(), 4),
            'Avg F1-Score': round(df_cls['F1-Score'].mean(), 4),
            'Std Precision': round(df_cls['Precision'].std(), 4),
            'Std Recall': round(df_cls['Recall'].std(), 4),
            'Std F1-Score': round(df_cls['F1-Score'].std(), 4),
            'Total Support': int(df_cls['Support'].sum()),
        })
    df_summary = pd.DataFrame(summary_rows)

    wb = Workbook()

    HEADER_FILL = PatternFill("solid", start_color="1F4E79")
    SUBHEADER_FILL = PatternFill("solid", start_color="2E75B6")
    ALT_FILL = PatternFill("solid", start_color="D6E4F0")
    WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_BODY = Font(name="Arial", bold=True, size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header_row(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = CENTER
            cell.border = thin_border

    def style_data_row(ws, row, cols, alt=False):
        fill = ALT_FILL if alt else WHITE_FILL
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = thin_border

    def write_df_to_sheet(ws, df, start_row=1, title=None):
        r = start_row
        if title:
            ws.cell(row=r, column=1, value=title).font = Font(name="Arial", bold=True, size=12, color="1F4E79")
            r += 1
        for c_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=r, column=c_idx, value=col_name)
        style_header_row(ws, r, len(df.columns))
        r += 1
        for row_idx, (_, row_data) in enumerate(df.iterrows()):
            alt = row_idx % 2 == 1
            for c_idx, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c_idx, value=val)
            style_data_row(ws, r, len(df.columns), alt=alt)
            r += 1
        return r

    # ---- Sheet 1: Summary ----
    ws_summary = wb.active
    ws_summary.title = "Summary"
    r = 1
    ws_summary.cell(row=r, column=1, value=f"Nested CV Results ({method_name}) — Summary")
    ws_summary.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_summary.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_summary.columns))
    r += 2

    write_df_to_sheet(ws_summary, df_summary, start_row=r)

    r += len(df_summary) + 3
    ws_summary.cell(row=r, column=1, value="Best Configuration").font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    r += 1
    config_dict = {
        "Best Sampling (down, up)": str(best_sampling),
        "n_estimators": best_hyperparams['n_estimators'],
        "max_depth": best_hyperparams['max_depth'],
        "min_samples_leaf": best_hyperparams['min_samples_leaf'],
        "Best Inner CV F1 (macro)": round(best_inner_score_overall, 4),
        "Num Features Selected": len(best_features) if best_features is not None else "N/A",
        "Selected Features": ", ".join(best_features) if best_features is not None else "N/A",
    }
    for key, val in config_dict.items():
        ws_summary.cell(row=r, column=1, value=key).font = BOLD_BODY
        ws_summary.cell(row=r, column=1).fill = ALT_FILL
        ws_summary.cell(row=r, column=2, value=val).font = BODY_FONT
        ws_summary.cell(row=r, column=2).fill = WHITE_FILL
        r += 1

    for col in range(1, len(df_summary.columns) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18
    ws_summary.column_dimensions['A'].width = 22

    # ---- Sheet 2: Outer Folds ----
    ws_outer = wb.create_sheet("Outer Folds")
    r = 1
    ws_outer.cell(row=r, column=1, value="Outer CV — Per-Fold Per-Class Metrics")
    ws_outer.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_outer.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_outer.columns))
    r += 2
    write_df_to_sheet(ws_outer, df_outer, start_row=r)
    for col in range(1, len(df_outer.columns) + 1):
        ws_outer.column_dimensions[get_column_letter(col)].width = 20
    ws_outer.column_dimensions['A'].width = 12

    # ---- Sheet 3: Inner Folds ----
    ws_inner = wb.create_sheet("Inner Folds")
    r = 1
    ws_inner.cell(row=r, column=1, value="Inner CV Metrics (per outer fold x sampling config x inner fold x class)")
    ws_inner.cell(row=r, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_inner.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df_inner.columns))
    r += 2
    write_df_to_sheet(ws_inner, df_inner, start_row=r)
    for col in range(1, len(df_inner.columns) + 1):
        ws_inner.column_dimensions[get_column_letter(col)].width = 22
    ws_inner.column_dimensions['A'].width = 12

    # ---- Sheet 4: Confusion Matrix ----
    ws_cm = wb.create_sheet("Confusion Matrix")
    ws_cm.cell(row=1, column=1, value="Overall Confusion Matrix (sum across outer folds)")
    ws_cm.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws_cm.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(unique_classes) + 2)

    ws_cm.cell(row=3, column=1, value="").font = BOLD_BODY
    ws_cm.cell(row=3, column=2, value="Predicted →").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws_cm.cell(row=3, column=2).fill = HEADER_FILL
    ws_cm.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(unique_classes) + 1)

    ws_cm.cell(row=4, column=1, value="Actual ↓").font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws_cm.cell(row=4, column=1).fill = SUBHEADER_FILL
    ws_cm.cell(row=4, column=1).alignment = CENTER

    for c_idx, cls in enumerate(unique_classes, 2):
        cell = ws_cm.cell(row=4, column=c_idx, value=f"Class {cls}")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = thin_border

    ws_cm.cell(row=4, column=len(unique_classes) + 2, value="Support").font = SUBHEADER_FONT
    ws_cm.cell(row=4, column=len(unique_classes) + 2).fill = SUBHEADER_FILL
    ws_cm.cell(row=4, column=len(unique_classes) + 2).alignment = CENTER

    for r_idx, (cls, cm_row) in enumerate(zip(unique_classes, overall_cm)):
        row = 5 + r_idx
        alt = r_idx % 2 == 1
        fill = ALT_FILL if alt else WHITE_FILL

        label_cell = ws_cm.cell(row=row, column=1, value=f"Class {cls}")
        label_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        label_cell.fill = SUBHEADER_FILL
        label_cell.alignment = CENTER

        for c_idx, val in enumerate(cm_row, 2):
            cell = ws_cm.cell(row=row, column=c_idx, value=int(val))
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = CENTER
            cell.border = thin_border

        ws_cm.cell(row=row, column=len(unique_classes) + 2, value=int(np.sum(cm_row))).font = BOLD_BODY
        ws_cm.cell(row=row, column=len(unique_classes) + 2).fill = fill
        ws_cm.cell(row=row, column=len(unique_classes) + 2).alignment = CENTER

    for col in range(1, len(unique_classes) + 3):
        ws_cm.column_dimensions[get_column_letter(col)].width = 14

    output_path = f"BLGPA_results_all_classes_fixed_{method_name}.xlsx"
    wb.save(output_path)
    print(f"\n[SUCCESS] Results for {method_name} saved to: {output_path}")

print("\n=======================================")
print(" ALL EXPERIMENTS COMPLETED SUCCESSFULLY")
print("=======================================")
